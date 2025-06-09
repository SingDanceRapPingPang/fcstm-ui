from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pyfcstm.model import State, CompositeState, PseudoState, Statechart

def export_statechart_to_excel(statechart: Statechart, file_path: str) -> bool:
    """
    将状态机信息导出为Excel文档
    
    Args:
        statechart: 状态机对象
        file_path: 导出文件路径
        
    Returns:
        bool: 是否导出成功
    """
    try:
        # 创建工作簿
        wb = Workbook()
        
        # 创建状态工作表
        states_sheet = wb.active
        states_sheet.title = "States"
        
        # 设置状态表头
        headers = ["状态名称", "状态类型", "状态描述", "Min时间锁", "Max时间锁", "Entry动作", "During动作", "Exit动作"]
        for col, header in enumerate(headers, 1):
            cell = states_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 添加状态信息
        row = 2
        for state in statechart.states:
            states_sheet.cell(row=row, column=1).value = state.name
            if isinstance(state, CompositeState):
                states_sheet.cell(row=row, column=2).value = "Composite"
            elif isinstance(state, PseudoState):
                states_sheet.cell(row=row, column=2).value = "Pseudo"
            else:
                states_sheet.cell(row=row, column=2).value = "Normal"
            states_sheet.cell(row=row, column=3).value = state.description if state.description else ""
            states_sheet.cell(row=row, column=4).value = state.min_time_lock if state.min_time_lock else ""
            states_sheet.cell(row=row, column=5).value = state.max_time_lock if state.max_time_lock else ""
            states_sheet.cell(row=row, column=6).value = state.on_entry if state.on_entry else ""
            states_sheet.cell(row=row, column=7).value = state.on_during if state.on_during else ""
            states_sheet.cell(row=row, column=8).value = state.on_exit if state.on_exit else ""
            row += 1
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            states_sheet.column_dimensions[chr(64 + col)].width = 15
        
        # 创建事件工作表
        events_sheet = wb.create_sheet("Events")
        
        # 设置事件表头
        headers = ["事件名称", "事件条件"]
        for col, header in enumerate(headers, 1):
            cell = events_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 添加事件信息
        row = 2
        events = set()
        for transition in statechart.transitions:
            if transition.event not in events:
                events.add(transition.event)
                events_sheet.cell(row=row, column=1).value = transition.event.name
                events_sheet.cell(row=row, column=2).value = transition.event.guard if transition.event.guard else ""
                row += 1
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            events_sheet.column_dimensions[chr(64 + col)].width = 15
        
        # 创建迁移工作表
        transitions_sheet = wb.create_sheet("Transitions")
        
        # 设置迁移表头
        headers = ["起始状态", "目标状态", "触发事件"]
        for col, header in enumerate(headers, 1):
            cell = transitions_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 添加迁移信息
        row = 2
        for transition in statechart.transitions:
            transitions_sheet.cell(row=row, column=1).value = transition.src_state.name
            transitions_sheet.cell(row=row, column=2).value = transition.dst_state.name
            transitions_sheet.cell(row=row, column=3).value = transition.event.name
            row += 1
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            transitions_sheet.column_dimensions[chr(64 + col)].width = 15
        
        # 保存工作簿
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"导出Excel文档时发生错误：{str(e)}")
        return False
